# Copyright © 2026 Nhat Viet Industrial Co., Ltd.
# All rights reserved.
#
# This software is proprietary and confidential.
# Unauthorized copying, modification, or distribution is strictly prohibited.

# app/report/copy_paste.py
"""
Module: Template Processor
Task: Process templates and replace placeholders in CSV/XLSX files
Description: Replaces placeholders in CSV/XLSX templates and generates report files

Functions:
- process_template(tpl_path, save_path, record): Process template and save result
  Detects file type (CSV/XLSX) and performs text replacements
  Returns dict: {"saved_path": path, "ext": extension}

Tác vụ: Xử lý và thay thế placeholder trong file CSV/XLSX
Mô tả: Thay thế placeholder trong template CSV/XLSX và tạo file report
"""

from typing import Dict
import os
import re

ph_re = re.compile(r"\[([^\[\]]+)\]")

def try_int(s):
    try:
        return int(s)
    except Exception:
        return None

def build_header_map(r) -> Dict[str, str]:
    m = {}
    m["record_id"] = r.record_id
    m["record_name"] = r.record_name
    m["source_pdf"] = r.source_pdf
    m["saved_date"] = r.saved_date or r.header.get("saved_date", "")
    m["work_name"] = r.work_name or r.header.get("work_name", "")
    m["work_number"] = r.work_number or r.header.get("work_number", "")
    m["operator"] = r.operator or r.header.get("operator", "")
    m["comments"] = r.comments or r.header.get("comments", "")
    m["total_crosswise_length_raw"] = r.header.get("total_crosswise_length_raw", r.header.get("total_crosswise_length", ""))
    m["crosswise_total_mm"] = r.crosswise_total_mm if getattr(r, "crosswise_total_mm", None) is not None else ""
    m["crosswise_segment_mm"] = r.header.get("crosswise_segment_mm", "")
    m["crosswise_segments_count"] = r.header.get("crosswise_segments_count", "")
    m["total_lengthwise_length_raw"] = r.header.get("total_lengthwise_length_raw", r.header.get("total_lengthwise_length", ""))
    m["lengthwise_total_mm"] = r.lengthwise_total_mm if getattr(r, "lengthwise_total_mm", None) is not None else ""
    m["lengthwise_segment_mm"] = r.header.get("lengthwise_segment_mm", "")
    m["lengthwise_segments_count"] = r.header.get("lengthwise_segments_count", "")
    m["flatness_um"] = r.flatness_um if getattr(r, "flatness_um", None) is not None else ""
    dims = {}
    try:
        dims = r.dims() or {}
    except Exception:
        dims = {}
    for k, v in dims.items():
        m[k] = v
    return m

def resolve_placeholder(token: str, rec, header_map: Dict[str,str]) -> tuple[str, bool]:
    """
    Resolve a placeholder token. 
    
    Returns: (resolved_value, was_resolved)
    - If placeholder is recognized and resolved, returns (value, True)
    - If placeholder is not recognized (e.g., graph placeholder), returns (original_token, False)
    
    Giải quyết một token placeholder.
    
    Trả về: (resolved_value, was_resolved)
    - Nếu placeholder được nhận diện và giải quyết, trả về (value, True)
    - Nếu placeholder không được nhận diện (vd: graph placeholder), trả về (original_token, False)
    """
    t = token.strip()
    # MEAS:AXIS:ROW:COL
    if t.upper().startswith("MEAS:"):
        parts = t.split(":")
        if len(parts) == 4:
            axis = parts[1].upper()
            row = try_int(parts[2])
            col = try_int(parts[3])
            if axis in ("X", "Y") and row is not None and col is not None:
                v = rec.measured.values.get((axis, row, col), "")
                return ("" if v is None else str(v), True)
        return ("", True)
    # RES:Y:X
    if t.upper().startswith("RES:"):
        parts = t.split(":")
        if len(parts) == 3:
            y = try_int(parts[1])
            x = try_int(parts[2])
            if y is not None and x is not None:
                v = rec.result.values.get((y, x), "")
                tag = rec.result.tags.get((y, x), "")
                s = "" if v is None else str(v)
                if tag:
                    s = f"{s} ({tag})"
                return (s, True)
        return ("", True)
    # header keys (case-insensitive)
    if t in header_map:
        return ("" if header_map[t] is None else str(header_map[t]), True)
    lk = t.lower()
    for k in header_map:
        if k.lower() == lk:
            return ("" if header_map[k] is None else str(header_map[k]), True)
    # Unknown placeholder - leave it unchanged
    return (f"[{t}]", False)

def _replace_in_string(s: str, rec, header_map: Dict[str,str]) -> str:
    if not isinstance(s, str) or "[" not in s:
        return s
    def repl(m):
        resolved_value, was_resolved = resolve_placeholder(m.group(1), rec, header_map)
        return resolved_value
    return ph_re.sub(repl, s)

def process_template(tpl_path: str, save_path: str, rec) -> Dict[str, str]:
    """
    Process the template file at tpl_path, replace placeholders using data from rec,
    and save result to save_path.

    Returns {"saved_path": save_path, "ext": ext}
    
    Xử lý file template tại tpl_path, thay thế placeholder bằng dữ liệu từ rec,
    và lưu kết quả vào save_path.

    Trả về {"saved_path": save_path, "ext": ext}
    """
    if not os.path.exists(tpl_path):
        raise FileNotFoundError(tpl_path)

    _, ext = os.path.splitext(tpl_path)
    ext = ext.lower()

    header_map = build_header_map(rec)

    if ext == ".csv":
        import csv
        with open(tpl_path, "r", encoding="utf-8", newline="") as fr:
            reader = csv.reader(fr)
            rows = [row for row in reader]
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if isinstance(cell, str) and "[" in cell:
                    rows[i][j] = _replace_in_string(cell, rec, header_map)
        with open(save_path, "w", encoding="utf-8", newline="") as fw:
            writer = csv.writer(fw)
            writer.writerows(rows)
        return {"saved_path": save_path, "ext": ext}

    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise e
        wb = load_workbook(tpl_path)
        for ws in wb.worksheets:
            # iterate values and replace strings
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and "[" in val:
                        cell.value = _replace_in_string(val, rec, header_map)
        wb.save(save_path)
        return {"saved_path": save_path, "ext": ext}
    else:
        raise ValueError(f"Unsupported template extension: {ext}")
