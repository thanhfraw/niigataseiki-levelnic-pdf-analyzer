# Copyright © 2026 Nhat Viet Industrial Co., Ltd.
# All rights reserved.
#
# This software is proprietary and confidential.
# Unauthorized copying, modification, or distribution is strictly prohibited.

# app/report/insert.py
"""
Module: Excel Image Inserter
Task: Insert and auto-fit images into XLSX workbook cells
Description: Helper module to insert and auto-fit images into XLSX workbook cells

Public function:
- insert_multiple_images_into_xlsx(xlsx_path, image_list)
  Inserts multiple images at specified placeholder positions
  Images are auto-fitted to Excel columns (A..C) or scaled by percentage

Tác vụ: Chèn ảnh vào file XLSX
Mô tả: Module hỗ trợ chèn ảnh vào ô workbook và tự động điều chỉnh kích thước
"""

import os
import shutil
import sys
from typing import Optional, Dict, List, Tuple

from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import pixels_to_EMU


def _safe_print(msg):
    """
    Print while swallowing encoding errors on Windows consoles.
    
    In ra màn hình và xử lý lỗi encoding trên console Windows.
    """
    try:
        print(msg)
    except UnicodeEncodeError:
        try:
            sys.stdout.buffer.write((str(msg) + "\n").encode("utf-8", errors="replace"))
        except Exception:
            pass
    except Exception:
        pass

# ==================== Helper Functions ====================

def _col_width_to_pixels(width):
    """
    Approximate Excel column width -> pixels.
    Formula: pixels ≈ width * 7 + 5 (common approximation)
    
    Ước lượng độ rộng cột Excel -> pixels.
    Công thức: pixels ≈ width * 7 + 5 (ước lượng thông dụng)
    """
    try:
        w = float(width)
    except Exception:
        w = 8.43  # default Excel column width
    return int(w * 7 + 5)


def _pixels_to_cm(pixels, dpi=96):
    """
    Convert pixels to centimeters (assume 96 DPI).
    
    Chuyển đổi pixels sang centimeters (giả định 96 DPI).
    """
    return pixels / float(dpi) * 2.54


def _fit_image_width_to_cm(img, target_width_cm, dpi=96):
    """
    Scale image to fit target width in centimeters.
    Preserves aspect ratio.
    
    Args:
        img: openpyxl.drawing.image.Image object
        target_width_cm: target width in centimeters
        dpi: DPI (default 96 for screen)
    
    Điều chỉnh kích thước ảnh để vừa với chiều rộng mục tiêu tính bằng centimeters.
    Giữ nguyên tỷ lệ khung hình.
    
    Tham số:
        img: đối tượng openpyxl.drawing.image.Image
        target_width_cm: chiều rộng mục tiêu tính bằng centimeters
        dpi: DPI (mặc định 96 cho màn hình)
    """
    try:
        orig_w_px = float(img.width)
        orig_h_px = float(img.height)
    except Exception:
        return img
    
    if orig_w_px <= 0:
        return img
    
    # Convert target width from cm to pixels
    target_width_px = target_width_cm / 2.54 * dpi
    scale = target_width_px / orig_w_px
    
    img.width = int(orig_w_px * scale)
    img.height = int(orig_h_px * scale)
    
    _safe_print(f"  Image scaled to {target_width_cm:.1f}cm width: {orig_w_px:.0f}x{orig_h_px:.0f}px -> {img.width:.0f}x{img.height:.0f}px (scale: {scale:.2f})")
    return img


def _fit_image_width_to_columns(ws, img, start_col="A", end_col="C"):
    """
    
    Điều chỉnh kích thước ảnh để vừa với tổng chiều rộng các cột từ start_col..end_col.
    Giữ nguyên tỷ lệ khung hình.
    
    Tham số:
        ws: worksheet openpyxl
        img: đối tượng openpyxl.drawing.image.Image
        start_col: chữ cái cột bắt đầu (mặc định "A")
        end_col: chữ cái cột kết thúc (mặc định "C")
    Scale image to fit combined width of columns start_col..end_col.
    Preserves aspect ratio.
    
    Args:
        ws: openpyxl worksheet
        img: openpyxl.drawing.image.Image object
        start_col: start column letter (default "A")
        end_col: end column letter (default "C")
    """
    from openpyxl.utils import column_index_from_string, get_column_letter
    
    CM_PER_PIXEL = 2.54 / 96.0
    
    # Convert column letters to indices
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    if end_idx < start_idx:
        start_idx, end_idx = end_idx, start_idx
    
    # Sum column widths in pixels
    total_pixels = 0
    for ci in range(start_idx, end_idx + 1):
        col_letter = get_column_letter(ci)
        col_dim = ws.column_dimensions.get(col_letter)
        width_val = 8.43  # default
        if col_dim is not None and getattr(col_dim, "width", None) is not None:
            width_val = col_dim.width
        px = _col_width_to_pixels(width_val)
        total_pixels += px
    
    # Get original image dimensions
    try:
        orig_w_px = float(img.width)
        orig_h_px = float(img.height)
    except Exception:
        return img  # can't scale if dimensions invalid
    
    # Calculate scale factor
    target_width_cm = _pixels_to_cm(total_pixels)
    orig_w_cm = orig_w_px * CM_PER_PIXEL
    
    if orig_w_cm <= 0:
        return img
    
    scale = target_width_cm / orig_w_cm
    
    # Apply scale
    img.width = int(orig_w_px * scale)
    img.height = int(orig_h_px * scale)
    
    _safe_print(f"  Image scaled to fit columns {start_col}..{end_col}: {orig_w_px:.0f}x{orig_h_px:.0f}px -> {img.width:.0f}x{img.height:.0f}px (scale: {scale:.2f})")
    
    return img


def _anchor_with_offset(cell, img, dx_px: int = 0, dy_px: int = 0) -> OneCellAnchor:
    """
    Create an anchor at the placeholder cell with pixel offsets.
    
    Tạo anchor tại ô placeholder với offset tính bằng pixel.
    """
    col_off_emu = int(pixels_to_EMU(max(0, dx_px)))
    row_off_emu = int(pixels_to_EMU(max(0, dy_px)))
    marker = AnchorMarker(col=cell.column - 1, colOff=col_off_emu, row=cell.row - 1, rowOff=row_off_emu)
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    ext = XDRPositiveSize2D(int(pixels_to_EMU(img.width)), int(pixels_to_EMU(img.height)))
    return OneCellAnchor(_from=marker, ext=ext)


# ==================== Main Insert Function ====================

def insert_multiple_images_into_xlsx(
    xlsx_path: str,
    images: List[Tuple[str, str]],
    copy_outputs_to_graph: bool = False,
    graph_dir: Optional[str] = None,
    scale_percent: float = 100.0,
    fit_columns: Optional[Tuple[str, str]] = None,
    target_width_cm: Optional[float] = None,
    offset_x_px: int = 50,
) -> Dict[str, List[Dict]]:
    """
    Insert images into Excel template.
    
    Args:
        xlsx_path: Path to XLSX file
        images: List of (image_path, placeholder) tuples
        copy_outputs_to_graph: Copy images to graph_dir (unused)
        graph_dir: Directory to copy images to (unused)
        scale_percent: Scale percentage (100 = full size). Ignored if fit_columns or target_width_cm is set.
        fit_columns: Auto-fit image to column range (e.g., ("A", "C")). Priority: target_width_cm > fit_columns
        target_width_cm: Target width in centimeters (e.g., 20 for A4). Priority: target_width_cm > fit_columns
    
    Returns: {"placed": [{"placeholder": placeholder, "cell": cell}, ...]}
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")
    
    for img_path, _ in images:
        if not os.path.exists(img_path):
            raise FileNotFoundError(f"Image not found: {img_path}")
    
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")
    
    wb = load_workbook(xlsx_path)
    placed_list = []
    
    try:
        for img_path, placeholder in images:
            # Insert original image (no scaling/resizing)
            img_to_insert = img_path
            
            # Find placeholder and insert image
            found = False
            for ws in wb.worksheets:
                if found:
                    break
                for row in ws.iter_rows():
                    if found:
                        break
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.strip() == placeholder:
                            cell.value = None
                            img = OpenpyxlImage(img_to_insert)
                            
                            # Auto-fit image size: priority is target_width_cm > fit_columns
                            if target_width_cm is not None:
                                try:
                                    _fit_image_width_to_cm(img, target_width_cm)
                                except Exception as e:
                                    _safe_print(f"  Width fit warning: {e}, using original size")
                            elif fit_columns:
                                try:
                                    _fit_image_width_to_columns(ws, img, fit_columns[0], fit_columns[1])
                                except Exception as e:
                                    _safe_print(f"  Column fit warning: {e}, using original size")
                            
                            # Anchor with a simple horizontal offset (applies to all images)
                            try:
                                if offset_x_px != 0:
                                    img.anchor = _anchor_with_offset(cell, img, dx_px=offset_x_px, dy_px=0)
                                    ws.add_image(img)
                                else:
                                    ws.add_image(img, cell.coordinate)
                            except Exception as e:
                                _safe_print(f"  Offset warning: {e}, using cell anchor")
                                ws.add_image(img, cell.coordinate)
                            placed_list.append({
                                "placeholder": placeholder,
                                "cell": cell.coordinate,
                                "sheet": ws.title
                            })
                            _safe_print(f"  Inserted {placeholder} at {cell.coordinate}")
                            found = True
                            break
        
        # Save workbook
        wb.save(xlsx_path)
        _safe_print(f"File saved: {xlsx_path}")
        
        return {"placed": placed_list}
    
    finally:
        try:
            wb.close()
        except:
            pass


def insert_single_image_into_xlsx(
    xlsx_path: str,
    image_path: str,
    placeholder: str,
    copy_outputs_to_graph: bool = False,
    graph_dir: Optional[str] = None,
    scale_percent: float = 100.0,
) -> Dict[str, str]:
    """Legacy function - calls insert_multiple_images_into_xlsx"""
    result = insert_multiple_images_into_xlsx(
        xlsx_path,
        [(image_path, placeholder)],
        copy_outputs_to_graph,
        graph_dir,
        scale_percent
    )
    if result["placed"]:
        placed = result["placed"][0]
        return {
            "placed_sheet": placed["sheet"],
            "placed_cell": placed["cell"]
        }
    raise ValueError(f"Placeholder {placeholder} not found")


